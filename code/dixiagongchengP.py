import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from statsmodels.graphics.tsaplots import plot_acf
from tsfresh import extract_features
from tsfresh.utilities.dataframe_functions import impute 
import tsfresh
from sklearn.feature_selection import SelectKBest, f_classif
from sklearn.feature_selection import SelectKBest, f_classif
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import accuracy_score
from sklearn.metrics import classification_report
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import MinMaxScaler  # 朴素贝叶斯使用MinMax缩放
from sklearn.naive_bayes import GaussianNB
from sklearn.metrics import classification_report, accuracy_score, confusion_matrix
import seaborn as sns
import matplotlib.pyplot as plt
import joblib
from openpyxl import load_workbook
import os
'''
1.加入原始  的resource_filter，因为有的  在同一个时间段有多列，造成无法对齐的情况
2.取消简单粗暴直接删除时间重复项，改为分组排序，取第一行
3.加入地铁线路运营商、线路遍历循环
'''

jizhan_to_remove=[]
def resource_filter(df,tel,jizhan_names):
    df = df[df['  '].str.startswith("%s"%tel)]        #控制只要哪个运营商的
    df = df[~df["  "].isin(jizhan_names)]
    return(df)



def  duiqi2(df_jizhan,complete_dates,jizhan_name,jizhan_class):
# 2. 将数据对齐到完整的时间序列索引
    df_jizhan.set_index('时间', inplace=True)  # 将日期列设为索引
    is_duplicate=df_jizhan.index.duplicated(keep=False)
     #= if_double # keep=False 标记所有重复项
    #df_jizhan.set_option('display.max_columns', None)
    #print(is_duplicate)
    df_jizhan=df_jizhan.drop_duplicates()  #去重
    #print(df_jizhan)
    # 先排序，再 groupby + first,去掉有多个时间段的数据，去重。
    df_jizhan = df_jizhan.sort_values("   ").groupby("时间").first().reset_index()
    df_jizhan.set_index('时间', inplace=True)  # 将日期列设为索引

    data_aligned = df_jizhan.reindex(complete_dates)  # 对齐到完整时间索引
    #data_aligned['标签']=data_aligned.loc['2025-04-25 14:00:00']['标签']   #用这个标签填充整个标签
    data_aligned['标签']=jizhan_class
    #data_aligned['  ']=data_aligned.loc['2025-04-25 14:00:00']['  ']
    #print(jizhan_name)
    data_aligned['  ']=jizhan_name
    data_aligned=data_aligned.fillna(0)
    data_aligned=data_aligned.reset_index()
    #print(data_aligned)
    return(data_aligned)

def excel_fenleijieguo(jizhan_names,excel_filename):
    #输出正确的分类
    excel_path = excel_filename
    new_data=jizhan_names
    # 将列表转换为 DataFrame（确保列名与原始文件一致）
    new_df = pd.DataFrame(new_data, columns=["Name"])
     
    # 2. 检查文件是否存在
    if os.path.exists(excel_path):
        # 文件存在 → 追加模式
        with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a") as writer:
            # 加载现有工作簿
            book = load_workbook(excel_path)
            writer.book = book  # 在 with 块内赋值（旧版 pandas 可能不支持）
            
            # 获取原始数据的最后一行
            sheet_name = book.sheetnames[0]  # 默认操作第一个 sheet
            original_df = pd.read_excel(excel_path, sheet_name=sheet_name)
            startrow = original_df.shape[0]  # 原始数据的行数
            
            # 追加数据（跳过标题行）
            new_df.to_excel(
                writer,
                sheet_name=sheet_name,
                startrow=startrow + 1,  # +1 跳过标题行
                index=False,
                header=False  # 不重复写入列名
            )
    else:
        # 文件不存在 → 创建新文件并写入数据（包含标题行）
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            new_df.to_excel(writer, sheet_name="Sheet1", index=False, header=True)
     
    print(f"数据已成功写入 {excel_path}")

def features_check(df1,df2):
    only_in_df1 = df1.columns.difference(df2.columns)
    only_in_df2 = df2.columns.difference(df1.columns)
     
    print("仅在 df1 中的列:", list(only_in_df1))
    print("仅在 df2 中的列:", list(only_in_df2))

    # 检查列名是否完全相同（包括顺序）
    columns_match = df1.columns.equals(df2.columns)
    print("列名和顺序是否完全一致？", columns_match)  # False


        # 检查列名是否相同（不考虑顺序）
    set1 = set(df1.columns)
    set2 = set(df2.columns)
    
    if set1 != set2:
        print ("不同，无法比较顺序")
    
    # 找出顺序不同的列
    mismatched_order = [col for i, col in enumerate(df1.columns) if col != df2.columns[i]]
    
    
    print("顺序不同的列",mismatched_order)
    # 按照 df1 的列顺序重新排列 df2
    df2 = df2.reindex(columns=df1.columns)
    return(df2)
        
    
 


def features2(df,y,yangben_name,number_of_ditie,tel_company):
    #extract_settings=ComprehensiveFCparameters()
    del df['标签']
    features = extract_features(df, 
        column_id='  ', 
        column_sort='index',
        #column_value='   ',
        #default_fc_parameters=extract_settings,
        impute_function=impute
        )
    #features.index.name='  '
    print(features)
  


    features.to_excel('output3.2.xlsx',index=True)
    #压缩维度
    '''
    x_imputed=impute(features)
    
    df.set_index('  ',inplace=True)
'''
    y['标签']=y['标签'].astype(bool)
    y.set_index('  ',inplace=True)
    print(y)

    '''
    print(df['标签'].unique())
    features['label']=[0,0,0,1,1,1,1,1]
    #y=pd.Series(y['标签'].values,index=y['  '].values)
    #print(y)
    x_selected=tsfresh.select_features(features,
        features['label'],
        fdr_level=0.01,
        feature_importance='f_classif',
        ml_task='classification'
        )
    print(x_selected)
    '''

    #download_robot_execution_failures()
    #timeseries, y = load_robot_execution_failures()
    #print(y)

    selector = SelectKBest(score_func=f_classif, k=20)  # 先获取所有特征的评分
    x_selected=selector.fit(features,y['标签'].values)

    #X_train, X_test, y_train, y_test = train_test_split(features, y['标签'].values, test_size=0.2, random_state=42)


    #导出
    results = pd.DataFrame({
    'Feature': features.columns,
    'Score': selector.scores_,
    'P-value': selector.pvalues_,
    'Selected': selector.get_support()
    })

    # 按评分降序排序
    results = results.sort_values('Score', ascending=False)
 
# 导出到Excel
    #x_selected.shape.to_excel('123.xlsx',index=False)
    results.to_excel('feature_selection_results2.xlsx', index=False)




    X_train, X_test, y_train, y_test = train_test_split(features, y['标签'].values, test_size=0.99, random_state=80)
    #X_test=features
    '''
    pipeline = Pipeline([
    ('selector', SelectKBest(score_func=f_classif, k=20)),  # 选择20个最佳特征
    ('scaler', MinMaxScaler(feature_range=(0, 1))),  # 将特征缩放到[0,1]区间
    ('gnb', GaussianNB())  # 高斯朴素贝叶斯
    ])
    print("开始训练朴素贝叶斯模型...")
    pipeline.fit(X_train, y_train)
    '''
    #读取模型
    pipeline = joblib.load('gaussianNB小时数据分类7.joblib') #读取pipeline
    xunlianji_data=joblib.load('1号线的训练集特征矩阵7.joblib') 


    #验证训练集和测试集的维度
    X_test=features_check(xunlianji_data,X_test) 

    X_test.to_excel("%s的测试集集特征矩阵.xlsx"%yangben_name)  #保存特征矩阵


    # 5. 评估模型
    y_pred = pipeline.predict(X_test)
    y_proba = pipeline.predict_proba(X_test)#[:, 1]  # 获取正类的预测概率
    #y_pred = (y_proba > 0.99).astype(int)
    print("\n分类报告:")
    print(classification_report(y_test, y_pred))
    #print(classification_report(y['标签'].values,y_pred))
    print(f"准确率: {accuracy_score(y_test, y_pred):.4f}")
    #print(f"准确率: {accuracy_score(y['标签'].values,y_pred):.4f}")


    # 6. 查看被选中的特征
    selected_features = X_train.columns[pipeline.named_steps['selector'].get_support()]
    print("\n被选中的特征:")
    print(selected_features)

    # 找出分类错误的项
    errors = []
    for i in range(len(y_test)):
        if y_test[i] != y_pred[i]:
            errors.append((i, y_test[i], y_pred[i],X_test.index[i],y_proba[i]))
    #print(X_test.index)
    # 显示分类错误的项
    print("\n分类错误的项:")
    jizhan_names=[]
    for index, true_label, pred_label,name_error,proba in errors:
        print(f"索引: {index}, 真实类别: {true_label}, 预测类别: {pred_label},名称：{name_error},概率评分:{proba}")
        jizhan_names.append(name_error)
    #输出结果
    excel_fenleijieguo(jizhan_names,excel_filename='%s%s负项分类结果.xlsx'%(number_of_ditie,tel_company))

        # 找出分类正确的项
    errors = []
    for i in range(len(y_test)):
        if y_test[i] == y_pred[i]:
            errors.append((i, y_test[i], y_pred[i],X_test.index[i],y_proba[i]))
    #print(X_test.index)
    # 显示分类错误的项
    print("\n分类正确的项:")
    jizhan_names=[]
    for index, true_label, pred_label,name_error,proba in errors:
        print(f"索引: {index}, 真实类别: {true_label}, 预测类别: {pred_label},名称：{name_error},概率评分:{proba}")
        jizhan_names.append(name_error)
    #输出结果
    excel_fenleijieguo(jizhan_names,excel_filename='%s%s正项分类结果.xlsx'%(number_of_ditie,tel_company))


    



    # 7. 可视化混淆矩阵
    plt.figure(figsize=(8, 6))
    cm = confusion_matrix(y_test, y_pred)
    sns.heatmap(cm, annot=True, fmt='d', cmap='Blues', 
                xticklabels=np.unique(y), yticklabels=np.unique(y))
    plt.xlabel('预测标签')
    plt.ylabel('真实标签')
    plt.title('混淆矩阵')
    #plt.show()

    # 8. 可视化特征重要性（基于SelectKBest的得分）
    plt.figure(figsize=(10, 6))
    scores = pipeline.named_steps['selector'].scores_[pipeline.named_steps['selector'].get_support()]
    sorted_idx = np.argsort(scores)
    plt.barh(range(len(sorted_idx)), scores[sorted_idx], align='center')
    plt.yticks(range(len(sorted_idx)), selected_features[sorted_idx])
    plt.xlabel('ANOVA F-value分数')
    plt.title('SelectKBest选中的特征重要性')
    plt.tight_layout()
    #plt.show()



def smooth_zero_fill2(series, max_zero_len=2):
    """
    平滑填充零值（支持连续1-2个零值）
    :param series: 输入序列
    :param max_zero_len: 最大处理零值长度
    :return: 处理后的序列
    """
    filled = series.copy().astype(float)  # 转为浮点型便于插值
    is_zero = filled == 0
    groups = is_zero.ne(is_zero.shift()).cumsum()
    zero_groups = groups[is_zero].unique()

    for zg in zero_groups:
        mask = groups == zg
        start, end = mask.idxmax(), mask.index[mask].max()
        length = end - start + 1
        
        if length > max_zero_len:
            continue

        # 寻找有效值边界
        prev_valid = filled[:start][filled[:start] != 0].iloc[-1] if not filled[:start].empty else None
        next_valid = filled[end+1:][filled[end+1:] != 0].iloc[0] if not filled[end+1:].empty else None

        # 边界情况处理
        if prev_valid is None and next_valid is None:
            continue  # 全零片段不处理
        elif prev_valid is None:
            filled[mask] = next_valid
        elif next_valid is None:
            filled[mask] = prev_valid
        else:
            # 线性插值计算
            positions = np.linspace(0, 1, length, endpoint=False)
            filled[mask] = prev_valid + (next_valid - prev_valid) * positions

    return filled.round(2)  # 保留两位小数







